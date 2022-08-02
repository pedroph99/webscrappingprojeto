
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import sys
"""""
moeda= 'dolar'
url='https://br.investing.com/currencies/streaming-forex-rates-majors'
getmoeda=requests.get(url)
getmoeda2=getmoeda.content

htmlmoeda=BeautifulSoup(getmoeda2, 'html.parser')
c=htmlmoeda.find('tbody')
print(len(c.find_all('tr')))
for x in range(len(c.find_all('tr'))):
    b=c.find_all('tr')[x].find('a').get('title')
    lista.append(b)
    preco=c.find_all('tr')[x].find_all('td')[2].text
    preco2=preco.replace(".", "")
    preco3=preco2.replace(',', ".")
    lista2.append((float(preco3)))
"""    
lista=[]
lista2=[]
def cria_workbook(nome, lista1, lista2, lista3):
    wb= openpyxl.load_workbook('jucepe.xlsx')
    ws=wb[wb.sheetnames[0]]


    counter=1
    counter2=0
    while counter <= len(lista1)+1:
        actual_cell=('A%i'%(counter))
        if actual_cell != 'A1':
            ws[actual_cell]=lista1[counter2]
            counter2+=1
        else:
            ws[actual_cell]='cnpj'

        counter+=1

    counter=1
    counter2=0
    while counter <= len(lista2)+1:
        actual_cell=('B%i'%(counter))
        if actual_cell != 'B1':
            ws[actual_cell]=lista2[counter2]
            counter2+=1
        else:
            ws[actual_cell]='Situacao'

        counter+=1
    counter=1
    counter2=0
    while counter <= len(lista3)+1:
        actual_cell=('C%i'%(counter))
        if actual_cell != 'C1':
            ws[actual_cell]=lista3[counter2]
            counter2+=1
        else:
            ws[actual_cell]='EMPRESA'

        counter+=1


    wb.save('jucepe.xlsx')
