import openpyxl
import os

import teste
def analisa(arquivo):
    #for widgets in frame.winfo_children():
    #  widgets.destroy()
    try:
        os.chdir('./Automacoa')
    except:
        pass
    print(os.getcwd())

    wb= openpyxl.load_workbook('%s.xlsx' %(arquivo))

    listacnpj=wb.sheetnames

    ws=wb[listacnpj[0]]
    #Identifica o campo CNPJ
    trace_row=None
    trace_column=None

    for x in range(ws.max_column+1):
        for y in range(ws.max_row+1):
            try:
                d=ws.cell(row=y, column=x)
                if d.value.lower() == 'cnpj':
                    trace_row=y
                    trace_column=x

            except:
                pass

    #pega todos os CNPJS de maneira vertical. 
    lista_cnpj=[]
    current_row=int(trace_row) +1
    while current_row <= int(ws.max_row):
        d=ws.cell(row=current_row, column=trace_column)
        lista_cnpj.append(d.value)

        current_row+=1



    for x in range(ws.max_column+1):
        for y in range(ws.max_row+1):
            try:
                d=ws.cell(row=y, column=x)
                if d.value.lower() == 'empresa':
                    trace_row=y
                    trace_column=x

            except:
                pass

    return lista_cnpj
print(analisa('CNPJJ'))